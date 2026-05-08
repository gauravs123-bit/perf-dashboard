import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

df = pd.read_csv('/tmp/no_track_creatives.csv')
df = df.sort_values('spend_l7d', ascending=False).reset_index(drop=True)

col_map = {
    'app': 'App',
    'ad_creative': 'Creative Name',
    'campaign': 'Campaign',
    'ad_set': 'Ad Set',
    'category': 'Category',
    'gender': 'Gender',
    'creative_type': 'Creative Type',
    'production': 'Production',
    'team': 'Team',
    'launch_month': 'Launch Month',
    'spend_yd': 'Spend YD (₹)',
    'orders_yd': 'D0 Orders YD',
    'CAC_yd': 'D0 CAC YD (₹)',
    'spend_l3d': 'Spend L3D (₹)',
    'orders_l3d': 'D0 Orders L3D',
    'CAC_l3d': 'D0 CAC L3D (₹)',
    'spend_l7d': 'Spend L7D (₹)',
    'orders_l7d': 'D0 Orders L7D',
    'CAC_l7d': 'D0 CAC L7D (₹)',
}

cols = list(col_map.keys())
headers = list(col_map.values())

spend_cols = {'spend_yd', 'CAC_yd', 'spend_l3d', 'CAC_l3d', 'spend_l7d', 'CAC_l7d'}

DARK_BG = '1e1e2e'
ARIVU_FILL = 'D6D4F5'
NERCHUKO_FILL = 'F5D4DE'

wb = Workbook()

# ── Creatives sheet ──────────────────────────────────────────────
ws = wb.active
ws.title = 'Creatives'

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', start_color=DARK_BG, fgColor=DARK_BG)
arivu_fill = PatternFill('solid', start_color=ARIVU_FILL, fgColor=ARIVU_FILL)
nerchuko_fill = PatternFill('solid', start_color=NERCHUKO_FILL, fgColor=NERCHUKO_FILL)

for c_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

ws.freeze_panes = 'A2'

num_fmt = '#,##0'
data_font = Font(name='Arial', size=10)

for r_idx, row in df.iterrows():
    excel_row = r_idx + 2
    app_val = row['app']
    row_fill = arivu_fill if app_val == 'Arivu' else nerchuko_fill

    for c_idx, col in enumerate(cols, 1):
        val = row[col]
        if isinstance(val, float) and math.isnan(val):
            val = None
        cell = ws.cell(row=excel_row, column=c_idx, value=val)
        cell.font = data_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical='center')
        if col in spend_cols and val is not None:
            cell.number_format = num_fmt

# Totals row
totals_row = len(df) + 2
spend_idx = {col: cols.index(col) + 1 for col in cols}

totals_label_cell = ws.cell(row=totals_row, column=1, value='TOTAL')
totals_label_cell.font = Font(name='Arial', bold=True, size=10)

total_fill = PatternFill('solid', start_color='2E2E3E', fgColor='2E2E3E')
total_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)

sum_cols = ['spend_yd', 'orders_yd', 'spend_l3d', 'orders_l3d', 'spend_l7d', 'orders_l7d']
cac_pairs = {
    'CAC_yd': ('spend_yd', 'orders_yd'),
    'CAC_l3d': ('spend_l3d', 'orders_l3d'),
    'CAC_l7d': ('spend_l7d', 'orders_l7d'),
}

data_start = 2
data_end = len(df) + 1

for c_idx, col in enumerate(cols, 1):
    cell = ws.cell(row=totals_row, column=c_idx)
    cell.fill = total_fill
    cell.font = total_font
    cell.alignment = Alignment(vertical='center')
    col_letter = get_column_letter(c_idx)
    if col in sum_cols:
        cell.value = f'=SUM({col_letter}{data_start}:{col_letter}{data_end})'
        if col in spend_cols:
            cell.number_format = num_fmt
    elif col in cac_pairs:
        spend_col, orders_col = cac_pairs[col]
        sc = get_column_letter(cols.index(spend_col) + 1)
        oc = get_column_letter(cols.index(orders_col) + 1)
        cell.value = f'=IF({oc}{totals_row}>0,{sc}{totals_row}/{oc}{totals_row},"")'
        cell.number_format = num_fmt
    elif col == 'app':
        cell.value = 'TOTAL'

# Auto-fit column widths
for c_idx, col in enumerate(cols, 1):
    col_letter = get_column_letter(c_idx)
    header_len = len(headers[c_idx - 1])
    if col == 'ad_creative':
        max_w = 60
    else:
        max_w = max(
            header_len + 2,
            max(
                (len(str(v)) for v in df[col].dropna()),
                default=header_len
            ) + 2
        )
        max_w = min(max_w, 40)
    ws.column_dimensions[col_letter].width = max_w

ws.row_dimensions[1].height = 20

# ── Summary sheet ────────────────────────────────────────────────
ws2 = wb.create_sheet('Summary')

sum_headers = ['App', '# Creatives', 'Total Spend L7D', 'Avg D0 CAC L7D', 'Total D0 Orders L7D']
for c_idx, h in enumerate(sum_headers, 1):
    cell = ws2.cell(row=1, column=c_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws2.freeze_panes = 'A2'

apps = ['Arivu', 'Nerchuko']
app_fills = {'Arivu': arivu_fill, 'Nerchuko': nerchuko_fill}

# Column indices in Creatives sheet (1-based)
app_col_letter = get_column_letter(cols.index('app') + 1)
spend_l7d_letter = get_column_letter(cols.index('spend_l7d') + 1)
orders_l7d_letter = get_column_letter(cols.index('orders_l7d') + 1)
cac_l7d_letter = get_column_letter(cols.index('CAC_l7d') + 1)

creatives_range = f'Creatives!{app_col_letter}{data_start}:{app_col_letter}{data_end}'

for r_idx, app in enumerate(apps, 2):
    row_fill = app_fills[app]
    # App
    c = ws2.cell(row=r_idx, column=1, value=app)
    c.font = Font(name='Arial', bold=True, size=10)
    c.fill = row_fill

    # # Creatives
    c2 = ws2.cell(row=r_idx, column=2)
    c2.value = f'=COUNTIF({creatives_range},A{r_idx})'
    c2.font = Font(name='Arial', size=10)
    c2.fill = row_fill

    # Total Spend L7D
    spend_range = f'Creatives!{spend_l7d_letter}{data_start}:{spend_l7d_letter}{data_end}'
    c3 = ws2.cell(row=r_idx, column=3)
    c3.value = f'=SUMIF({creatives_range},A{r_idx},{spend_range})'
    c3.font = Font(name='Arial', size=10)
    c3.fill = row_fill
    c3.number_format = num_fmt

    # Total D0 Orders L7D
    orders_range = f'Creatives!{orders_l7d_letter}{data_start}:{orders_l7d_letter}{data_end}'
    c5 = ws2.cell(row=r_idx, column=5)
    c5.value = f'=SUMIF({creatives_range},A{r_idx},{orders_range})'
    c5.font = Font(name='Arial', size=10)
    c5.fill = row_fill

    # Avg D0 CAC L7D = Total Spend / Total Orders
    c4 = ws2.cell(row=r_idx, column=4)
    c4.value = f'=IF(E{r_idx}>0,C{r_idx}/E{r_idx},"")'
    c4.font = Font(name='Arial', size=10)
    c4.fill = row_fill
    c4.number_format = num_fmt

# Grand total row
grand_row = len(apps) + 2
gt_fill = PatternFill('solid', start_color='2E2E3E', fgColor='2E2E3E')
gt_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)

labels = ['TOTAL', None, None, None, None]
for c_idx in range(1, 6):
    cell = ws2.cell(row=grand_row, column=c_idx)
    cell.fill = gt_fill
    cell.font = gt_font
    cell.alignment = Alignment(vertical='center')
    col_letter_s = get_column_letter(c_idx)
    if c_idx == 1:
        cell.value = 'TOTAL'
    elif c_idx == 2:
        cell.value = f'=SUM({col_letter_s}2:{col_letter_s}{grand_row-1})'
    elif c_idx == 3:
        cell.value = f'=SUM(C2:C{grand_row-1})'
        cell.number_format = num_fmt
    elif c_idx == 5:
        cell.value = f'=SUM(E2:E{grand_row-1})'
    elif c_idx == 4:
        cell.value = f'=IF(E{grand_row}>0,C{grand_row}/E{grand_row},"")'
        cell.number_format = num_fmt

# Auto-fit Summary columns
sum_widths = [15, 14, 18, 16, 22]
for c_idx, w in enumerate(sum_widths, 1):
    ws2.column_dimensions[get_column_letter(c_idx)].width = w

ws2.row_dimensions[1].height = 20

out_path = '/Users/gauravsharma/Claude/perf_dashboard/no_track_creatives.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
